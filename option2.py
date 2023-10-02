import tkinter as tk
import openpyxl
from tkinter import messagebox
import re
from tkinter import ttk

class option2_class:
    def __init__(self,root):
        
        root.geometry("900x600")
        self.bg_image = tk.PhotoImage(file="bg1.png")
        self.label1 = tk.Label(root,image=self.bg_image).place(x = 0, y = 30, width = 1280, height = 700)

        # root.geometry("1280x700+0+0")
        root.option_add("*Font", "Verdana 10")
        root.grid_rowconfigure(2, minsize=20)
        root.title("Request a CallBack")
        label= tk.Label(root, text= "Request a CallBack", width=55, height=3, relief="solid", borderwidth=1, bg='teal', font=("Verdana", 20), fg="white")
        label.grid(row=0, column=0, columnspan=2)

        # create labels and input fields for the form
        first_name_label = tk.Label(root, text="First Name")
        first_name_label.grid(row=5, column=0)
        root.grid_rowconfigure(2, minsize=20)
        #first_name_label.pack()
        first_name_entry = tk.Entry(root, borderwidth=3, relief="groove", width= 40)
        first_name_entry.grid(row=5, column=1, pady=10)
        #first_name_entry.pack()

        last_name_label = tk.Label(root, text="Last Name")
        #last_name_label.pack()
        last_name_label.grid(row=7, column=0)
        last_name_entry = tk.Entry(root, borderwidth=3, relief="groove", width= 40)
        #last_name_entry.pack()
        last_name_entry.grid(row=7, column=1, pady=10)

        mobile_label = tk.Label(root, text="Mobile Number")
        #mobile_label.pack()
        mobile_label.grid(row=9, column=0)
        mobile_entry = tk.Entry(root, borderwidth=3, relief="groove", width= 40)
        mobile_entry.grid(row=9, column=1, pady=10)
        #mobile_entry.pack()

        email_label = tk.Label(root, text="Email Address")
        email_label.grid(row=11, column=0)
        #email_label.pack()
        email_entry = tk.Entry(root, borderwidth=3, relief="groove", width= 40)
        email_entry.grid(row=11, column=1, pady=10)
        #email_entry.pack()
   

        hospitals = ["Delhi", "Maharashtra", "Gujarat", "Punjab", "Rajasthan","Kerela"]
        # Create a label for the hospital name
        hospital_label = ttk.Label(root, text="Hospital")
        hospital_label.grid(row=13, column=0, pady=10)

        # Create a Combobox for the hospital name
        hospital_combobox = ttk.Combobox(root, values=hospitals, state="readonly")
        hospital_combobox.grid(row=13, column=1, pady=10)

        # Set the default value of the Combobox
        hospital_combobox.current(0)

        """hospital_label = tk.Label(root, text="Hospital Name")
        hospital_label.grid(row=13, column=0)
        #hospital_label.pack()
        hospital_entry = tk.Entry(root, borderwidth=3, relief="groove", width= 40)
        hospital_entry.grid(row=13, column=1, pady=10)"""
        #hospital_entry.pack()


        comments_label = tk.Label(root, text="Comments (Optional)")
        comments_label.grid(row=15, column=0)
        #comments_label.pack()
        comments_entry = tk.Text(root, height=3, width=60, borderwidth=3, relief="groove")
        comments_entry.grid(row=15, column=1, pady=25)
        #comments_entry.pack()

        # create a submit button
        submit_button = tk.Button(root, text="Submit", command=self.submit_callback)
        #submit_button.pack()
        submit_button.grid(row=30, column=1, pady=40)

                
    def submit_callback(self):
        # get the values entered by the user
        first_name = self.first_name_entry.get().strip()
        last_name = self.last_name_entry.get().strip()
        mobile = self.mobile_entry.get().strip()
        email = self.email_entry.get().strip()
        hospital = self.hospital_combobox.get().strip()
        comments = self.comments_entry.get("1.0", "end-1c")
        # get the text from the text widget
        
        # validate the form data
        # check that all required fields are filled in
        if not all([first_name, last_name, mobile, email]):
            error_message = "Please fill in all required fields"
            tk.messagebox.showerror("Error", error_message)
            return
        
        # check if mobile number is valid
        if not re.match(r'^\d{10}$', mobile):
            messagebox.showerror("Error", "Please enter a valid 10-digit mobile number.")
            return

        # check if email address is valid
        if not re.match(r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$', email):
            messagebox.showerror("Error", "Please enter a valid email address.")
            return
        
        # create a message with the form data
        message = [first_name, last_name, mobile, email, hospital, comments]
        
        # save the message to an Excel file
        wb = openpyxl.load_workbook('callback_requests.xlsx')
        ws = wb.active
        row = ws.max_row + 1
        for i, value in enumerate(message):
            ws.cell(row=row, column=i+1, value=value)
            wb.save('callback_requests.xlsx')
        
        # show a success message
        success_message = "Your call back request has been submitted"
        tk.messagebox.showinfo("Success", success_message)
        # close the root
        root.destroy()

        

if __name__ == "__main__":
    root = tk.Tk()
    obj = option2_class(root)
    root.mainloop()