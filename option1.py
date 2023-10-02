import tkinter as tk
import openpyxl
from tkinter import messagebox
import re
from tkinter import ttk
from tkcalendar import DateEntry

class option1_class:
    def __init__(self,root):
        
        root.geometry("900x600")
        self.bg_image = tk.PhotoImage(file="bg1.png")
        self.label1 = tk.Label(root,image=self.bg_image).place(x = 0, y = 30, width = 1280, height = 700)

        root.option_add("*Font", "Verdana 10")
        root.grid_rowconfigure(2, minsize=20)
        root.title("Book an Appointment")
        label= tk.Label(root, text= "Book an Appointment", width=55, height=3, relief="solid", borderwidth=1, bg='teal', font=("Verdana", 20), fg="white")
        label.grid(row=0, column=0, columnspan=2)

        #CREATE NAME
        name_label = tk.Label(root, text="Full Name")
        name_label.grid(row=5, column=0)
        root.grid_rowconfigure(2, minsize=20)
        name_entry = tk.Entry(root, borderwidth=3, relief="groove", width= 40)
        name_entry.grid(row=5, column=1, pady=10)

        #CREATE MOBILE NUMBER
        mobile_label = tk.Label(root, text="Mobile Number")
        mobile_label.grid(row=9, column=0)
        mobile_entry = tk.Entry(root, borderwidth=3, relief="groove", width= 40)
        mobile_entry.grid(row=9, column=1, pady=10)

        #CREATE EMAIL ID
        email_label = tk.Label(root, text="Email Address")
        email_label.grid(row=11, column=0)
        email_entry = tk.Entry(root, borderwidth=3, relief="groove", width= 40)
        email_entry.grid(row=11, column=1, pady=10)

        #CREATE HOSPITAL DROP DOWN LIST
        hospitals = ["Hospital A", "Hospital B", "Hospital C", "Hospital D", "Hospital E"]
        # Create a label for the hospital name
        hospital_label = ttk.Label(root, text="Hospital Name")
        hospital_label.grid(row=13, column=0, pady=10)
        # Create a Combobox for the hospital name
        hospital_combobox = ttk.Combobox(root, width=30, values=hospitals, state="readonly")
        hospital_combobox.grid(row=13, column=1, pady=10)
        # Set the default value of the Combobox
        #hospital_combobox.current(0)

        #CREATE DEPARTMENT DROP DOWN LIST
        department = ["Anesthesiology", "Cardiology", "Dermatology", "Emergency medicine", "Endocrinology", "Gastroenterology",
                    "General medicine", "General surgery", "Hematology", "Nephrology", "Neurology", "Obstetrics and gynecology", 
                    "Oncology", "Ophthalmology", "Orthopedics", "Otolaryngology", "Pediatrics", "Physical medicine and rehabilitation",
                    "Psychiatry", "Pulmonology", "Radiology", "Rheumatology", "Urology"]
        # Create a label for the department name
        department_label = ttk.Label(root, text="Department")
        department_label.grid(row=15, column=0, pady=10)
        # Create a Combobox for the department name
        department_combobox = ttk.Combobox(root, width=30, values=department, state="readonly")
        department_combobox.grid(row=15, column=1, pady=10)
        # Set the default value of the Combobox
        #hospital_combobox.current(0)

        #CREATE CALENDAR TO SELECT THE DATE
        date_label = tk.Label(root, text="Select a date:")
        date_label.grid(row=17, column=0)
        date_entry = DateEntry(root, width=15, background='darkblue', foreground='white', borderwidth=2)
        date_entry.grid(row=17, column=1, pady=10)

        #CREATE TIME SLOT DROP DOWN
        timeslot = ['09:00 AM', '10:00 AM', '11:00 AM', '12:00 PM', '01:00 PM', '03:00 PM', '04:00 PM', '05:00 PM', '07:00 PM', '08:00 PM']
        # Create a label for the time slot
        time_label = ttk.Label(root, text="Time Slot")
        time_label.grid(row=19, column=0, pady=10)
        # Create a Combobox for the time slot
        time_combobox = ttk.Combobox(root, width=30, values=timeslot, state="readonly")
        time_combobox.grid(row=19, column=1, pady=10)
        # Set the default value of the Combobox
        #time_combobox.current(0)

        # create a submit button
        submit_button = tk.Button(root, text="Submit", command=self.book_appointment)
        submit_button.grid(row=30, column=1, pady=40)
        
##TITLE
    def book_appointment(self):
        # get the values entered by the user
        full_name = self.name_entry.get().strip()
        mobile = self.mobile_entry.get().strip()
        email = self.email_entry.get().strip()
        hospital = self.hospital_combobox.get().strip()
        department= self.department_combobox.get().strip()
        date= self.date_entry.get_date()
        time = self.time_combobox.get().strip()
                
        # check that all required fields are filled in
        if not all([full_name, mobile, email, date]):
            error_message = "Please fill in all required fields"
            tk.messagebox.showerror("Error", error_message)
            return
        # check if mobile number is valid
        if not re.match(r'^\d{10}$', mobile):
            messagebox.showerror("Error", "Please enter a valid 10-digit mobile number.")
            return

        # check if email address is valid
        if not re.match(r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$', email):
            messagebox.showerror("Error", "Please enter a valid email address.")
            return
        
        # create a message with the form data
        message = [full_name, mobile, email, hospital, department, date, time]
        
        # save the message to an Excel file
        wb = openpyxl.load_workbook('appointment_list.xlsx')
        ws = wb.active
        row = ws.max_row + 1
        for i, value in enumerate(message):
            ws.cell(row=row, column=i+1, value=value)
            wb.save('appointment_list.xlsx')
        
        # show a success message
        success_message = "Your appointment request has been submitted. Details will be shared to the given mobile number."
        tk.messagebox.showinfo("Success", success_message)
        # close the root
        root.destroy()

        

if __name__ == "__main__":
    root = tk.Tk()
    obj = option1_class(root)
    root.mainloop()